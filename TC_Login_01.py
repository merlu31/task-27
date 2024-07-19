import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

from pageObjects.LoginPage import LoginPage


# Function to load test data from Excel
def load_test_data(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    data = []

    # Skip the header row (index 1)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append((row[0], row[1], row[2]))  # Include Test ID, username, and password

    return data


# Fixture to initialize and quit the WebDriver
@pytest.fixture()
def driver():
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver.maximize_window()
    driver.implicitly_wait(10)
    yield driver
    driver.close()
    driver.quit()


# Parameterize test with data from Excel
excelPath = "D:\\Automation\\Test\\pythonProject\\Excel\\test_data.xlsx"
@pytest.mark.parametrize("test_id, username, password",
                         load_test_data(excelPath))
def test_data_driven_login_test(driver, test_id, username, password):
    login_page = LoginPage(driver)
    login_page.open_page("https://opensource-demo.orangehrmlive.com")

    login_page.enter_username(username)
    login_page.enter_password(password)
    login_page.click_login_button()

    if "dashboard" in driver.current_url.lower():
        print(f"The user '{username}' is logged in successfully.")
        result = "passed"
    else:
        print(f"The user '{username}' failed to login.")
        result = "failed"

    # Write result back to Excel sheet
    write_test_result(test_id, result)

    assert "dashboard" in driver.current_url.lower(), f"Login failed for user '{username}'."


def write_test_result(test_id, result):
    file_name = excelPath
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    # Find the row for the current test ID
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == test_id:
            sheet.cell(row=row, column=4).value = result  # Write result in the fourth column
            break

    workbook.save(file_name)


if __name__ == "__main__":
    pytest.main()
